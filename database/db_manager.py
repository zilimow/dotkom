import sqlite3
import logging
from contextlib import contextmanager
from typing import List, Dict, Any, Optional
import pandas as pd
import io
from datetime import datetime
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

COLUMNS_EQUIPMENT = [
    "Бортовой номер",   # A (eq_board)
    "Тип техники",      # B (eq_type)
    "Модель",           # C (eq_model)
    "Серийный номер",   # D (eq_serial)
    "Год производства", # E (eq_year)
    "ДВС",              # F (eq_engine)
    "Номер двигателя",  # G (eq_engine_number)
    "Код"               # H (eq_code)
]

DB_FILE = "fleet.db"

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class DatabaseError(Exception):
    pass


@contextmanager
def get_connection_context():
    """Context manager for database connections with enabled cascade deletion."""
    conn = None
    try:
        conn = sqlite3.connect(DB_FILE)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("PRAGMA foreign_keys = ON;")
        yield conn
        conn.commit()
    except sqlite3.Error as e:
        if conn:
            conn.rollback()
        logger.error(f"Database error: {e}")
        raise DatabaseError(f"Database operation failed: {e}") from e
    finally:
        if conn:
            conn.close()



def init_db():
    """Initialize database with all required tables and indexes."""
    with get_connection_context() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS equipment (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                eq_board TEXT NOT NULL CHECK(trim(eq_board) != ''),
                eq_type TEXT NOT NULL CHECK(trim(eq_type) != ''),
                eq_model TEXT,
                eq_serial TEXT,
                eq_year TEXT,
                eq_engine TEXT,
                eq_engine_number TEXT,
                eq_code TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS work_orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                equipment_id INTEGER NOT NULL,          -- Связь по системному ID
                work_date DATE DEFAULT CURRENT_DATE,
                work_task TEXT,
                work_desc TEXT,
                work_hours REAL,
                time_start TEXT,
                time_end TEXT,
                work_executor TEXT,
                work_notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (equipment_id) REFERENCES equipment(id) ON DELETE CASCADE -- Ссылка на id
            )
        """)
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_work_orders_eq_id ON work_orders(equipment_id)")
        
               
        logger.info("Database initialized successfully")


def load_equipment():
    """Load all equipment from database."""
    with get_connection_context() as conn:
        df = pd.read_sql_query("SELECT * FROM equipment ORDER BY eq_board", conn)
        return df

def add_equipment(eq_board, eq_type, eq_model, eq_serial, eq_year, eq_engine, eq_engine_number, eq_code):
    """Add new equipment to database safely without erasing different types."""
    with get_connection_context() as conn:
        cursor = conn.cursor()
        # 🚨 ИСПРАВЛЕНИЕ: Безопасный INSERT. Если связка Борт+Модель+Тип уже есть — СУБД выдаст ошибку,
        # а не сотрет данные молча, как делал OR REPLACE. Валидация контролируется в Streamlit.
        cursor.execute("""
            INSERT INTO equipment (eq_board, eq_type, eq_model, eq_serial, eq_year, eq_engine, eq_engine_number, eq_code)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (eq_board, eq_type, eq_model, eq_serial, eq_year, eq_engine, eq_engine_number, eq_code))


def update_equipment(equipment_id, **kwargs):
    """Update existing equipment."""
    if not kwargs:
        return True
    
    set_clause = ', '.join([f"{key} = ?" for key in kwargs.keys()])
    query = f"UPDATE equipment SET {set_clause}, updated_at = CURRENT_TIMESTAMP WHERE id = ?"
    
    with get_connection_context() as conn:
        cursor = conn.cursor()
        values = list(kwargs.values()) + [equipment_id]
        cursor.execute(query, values)
        return cursor.rowcount > 0


def delete_equipment(equipment_id):
    """Delete equipment permanently."""
    with get_connection_context() as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM equipment WHERE id = ?", (equipment_id,))
        return cursor.rowcount > 0


def get_equipment_statistics():
    """Get statistics about equipment fleet."""
    with get_connection_context() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM equipment")
        total = cursor.fetchone()[0]      
        return {'total': total}

def load_work_orders() -> pd.DataFrame:
    """Загрузка полного журнала работ с автоматическим подтягиванием данных из паспорта техники."""
    with get_connection_context() as conn:
        query = """
            SELECT 
                w.work_date AS [eq_date],
                e.eq_board AS [eq_board],       -- Борт подтягивается по ID связи
                e.eq_type AS [eq_type],         -- Тип подтягивается по ID связи
                e.eq_model AS [eq_model],       -- Модель подтягивается по ID связи
                w.work_task AS [eq_task],
                w.work_desc AS [eq_desc],
                w.work_hours AS [eq_hours],
                w.time_start AS [time_start],
                w.time_end AS [time_end],
                w.work_executor AS [eq_executor],
                w.work_notes AS [eq_notes],
                w.id AS [id]
            FROM work_orders w
            LEFT JOIN equipment e ON w.equipment_id = e.id
            ORDER BY w.work_date DESC, w.id DESC
        """
        df = pd.read_sql_query(query, conn)
        return df

def add_work_order(
    equipment_id: int,
    work_date: str, 
    work_task: str,
    work_desc: str,
    work_hours: float,
    time_start: str,
    time_end: str,
    work_executor: str,
    work_notes: str,
):
    """Добавление новой записи в журнал ТОиР с сохранением даты в текстовом формате ДД.ММ.ГГГГ."""
    with get_connection_context() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO work_orders (
                equipment_id, 
                work_date, 
                work_task, 
                work_desc, 
                work_hours, 
                time_start, 
                time_end, 
                work_executor, 
                work_notes
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                equipment_id,
                work_date,  
                work_task,
                work_desc,
                work_hours,
                time_start,
                time_end,
                work_executor,
                work_notes,
            ),
        )

def update_work_order(
    record_id: int,
    work_date: str, 
    work_task: str,
    work_desc: str,
    work_hours: float,
    time_start: str,
    time_end: str,
    work_executor: str,
    work_notes: str,
):
    """Обновление существующей записи в журнале ТОиР по ее системному id."""
    with get_connection_context() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            UPDATE work_orders 
            SET 
                work_date = ?, 
                work_task = ?, 
                work_desc = ?, 
                work_hours = ?, 
                time_start = ?, 
                time_end = ?, 
                work_executor = ?, 
                work_notes = ?
            WHERE id = ?
        """,
            (
                work_date,  
                work_task,
                work_desc,
                work_hours,
                time_start,
                time_end,
                work_executor,
                work_notes,
                record_id, # Уникальный ID записи для условия WHERE
            ),
        )

def delete_work_order(record_id: int):
    """Удаление записи из журнала ТОиР по ее системному id."""
    with get_connection_context() as conn:
        cursor = conn.cursor()
        # Принудительно преобразуем к int и передаем кортеж (record_id,)
        cursor.execute(
            "DELETE FROM work_orders WHERE id = ?;",
            (int(record_id),)
        )
        
def generate_equipment_blank_template() -> io.BytesIO:
    """Генерация чистого Excel-шаблона с красивой жирной шапкой и готовой шириной столбцов."""
    output = io.BytesIO()
    
    df_blank = pd.DataFrame(columns=COLUMNS_EQUIPMENT)
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_blank.to_excel(writer, sheet_name="Техника", index=False)
        worksheet = writer.sheets["Техника"]
        
        # 1. Стилизация шапки (как в основном экспорте)
        font_header = Font(name="Calibri", size=13, bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Устанавливаем увеличенную высоту для строки заголовков
        worksheet.row_dimensions[1].height = 26
        
        # Применяем жирный шрифт и центрирование к каждой ячейке шапки
        for col_idx in range(1, len(COLUMNS_EQUIPMENT) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = center_alignment
        
        # 2. Выставляем фиксированную ширину ячеек
        column_widths = {
            "A": 22, "B": 28, "C": 22, "D": 25, 
            "E": 25, "F": 28, "G": 25, "H": 18
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

    output.seek(0)
    return output


def import_equipment_from_excel(uploaded_file) -> tuple[bool, str]:
    """Чтение Excel-файла, жесткая очистка от грязи и безопасный импорт в SQLite 
    БЕЗ ПЕРЕТИРАНИЯ существующих записей в базе данных.
    """
    try:
        # Шаг 1: Чтение файла (загружаем всё как текст)
        df = pd.read_excel(uploaded_file, sheet_name=0, dtype=str)
        
        if df.empty:
            return False, "Файл пуст. Заполните шаблон данными."
            
        # Шаг 2: Проверка структуры колонок
        missing_cols = [col for col in COLUMNS_EQUIPMENT if col not in df.columns]
        if missing_cols:
            return False, f"Неверная структура шаблона. Отсутствуют колонки: {', '.join(missing_cols)}"
        
        # Заменяем реальные NaN/None от pandas на пустые строки
        df = df.fillna("")
        
        # Удаляем строки, где вообще не заполнен Тип техники
        df = df[df["Тип техники"].astype(str).str.strip() != ""]
        
        if df.empty:
            return False, "В файле нет валидных строк (пропущен Тип техники)."

        # Очистка ячеек от мусора системных типов (.0, nan)
        bn_counter = 1
        for col in df.columns:
            df[col] = df[col].astype(str).str.strip()
            df[col] = df[col].replace({"None": "", "nan": "", "None.0": "", "nan.0": ""})
            df[col] = df[col].str.replace(r"\.0$", "", regex=True)

        # Обрабатываем пустые бортовые номера
        for index, row in df.iterrows():
            if row["Бортовой номер"] == "":
                df.at[index, "Бортовой number"] = f"б/н-{bn_counter}"
                bn_counter += 1

        # --- ГЛАВНЫЙ ФИЛЬТР ЗАЩИТЫ ОТ ПЕРЕТИРАНИЯ ---
        # Выкачиваем текущую базу, чтобы сравнить дубликаты прямо в памяти
        with get_connection_context() as conn:
            df_current = pd.read_sql_query("SELECT eq_board, eq_model, eq_type FROM equipment", conn)
        
        # Собираем множество существующих машин по правилу UNIQUE вашей БД (Бортовой + Модель + Тип)
        # Переводим в нижний регистр для точного сравнения без учета регистра
        existing_records = set()
        if not df_current.empty:
            for _, r in df_current.iterrows():
                key = (str(r['eq_board']).lower().strip(), 
                       str(r['eq_model']).lower().strip(), 
                       str(r['eq_type']).lower().strip())
                existing_records.add(key)

        inserted_count = 0
        skipped_count = 0

        # Шаг 4: Запись в базу данных
        with get_connection_context() as conn:
            cursor = conn.cursor()
            
            for _, row in df.iterrows():
                # Строим проверочный ключ для текущей строки из файла
                row_key = (row["Бортовой номер"].lower(), 
                           row["Модель"].lower(), 
                           row["Тип техники"].lower())
                
                # 🚨 ЗАЩИТА: Если такая связка уже есть в базе — пропускаем строку, ничего не перетирая!
                if row_key in existing_records:
                    skipped_count += 1
                    continue
                
                try:
                    # Использован чистый INSERT. Он не перезапишет строку в случае коллизии
                    cursor.execute("""
                        INSERT OR IGNORE INTO equipment (
                            eq_board, eq_type, eq_model, eq_serial, eq_year, 
                            eq_engine, eq_engine_number, eq_code
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        row["Бортовой номер"],   
                        row["Тип техники"],      
                        row["Модель"],           
                        row["Серийный номер"],   
                        row["Год производства"], 
                        row["ДВС"],              
                        row["Номер двигателя"],  
                        row["Код"]               
                    ))
                    if cursor.rowcount > 0:
                        inserted_count += 1
                    else:
                        skipped_count += 1
                except sqlite3.Error as e:
                    logger.warning(f"Ошибка импорта строки {row['Бортовой номер']}: {e}")
                    skipped_count += 1
                    continue
                    
        # Выдаем детальный развернутый отчет оператору РММ
        if inserted_count == 0:
            return True, f"Импорт завершен. Все машины ({skipped_count} шт.) из файла уже есть в базе. Никакие данные не были изменены."
            
        msg = f"Успешно добавлено новых машин: {inserted_count}."
        if skipped_count > 0:
            msg += f" Пропущено дубликатов: {skipped_count} шт. (база защищена от перетирания)."
            
        return True, msg
        
    except Exception as e:
        logger.error(f"Ошибка импорта: {e}")
        return False, f"Ошибка при обработке файла: {str(e)}"


def export_equipment_to_excel() -> io.BytesIO:
    """Выгрузка записей из таблицы equipment в Excel с полосами, жирным бортовым номером,
    увеличенной высотой строк, 10 пустыми строками и преднастройкой печати.
    """
    output = io.BytesIO()

    with get_connection_context() as conn:
        query = """
            SELECT 
                eq_board AS [Бортовой номер], 
                eq_type AS [Тип техники], 
                eq_model AS [Модель], 
                eq_serial AS [Серийный номер], 
                eq_year AS [Год производства], 
                eq_engine AS [ДВС], 
                eq_engine_number AS [Номер двигателя],
                eq_code AS [Код]
            FROM equipment 
            ORDER BY eq_type ASC, eq_board ASC
        """
        df_export = pd.read_sql_query(query, conn)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, sheet_name="Список_Техники", index=False)
        worksheet = writer.sheets["Список_Техники"]

        # 1. Настройка стилей текста, выравнивания и границ
        font_header = Font(name="Calibri", size=13, bold=True)
        font_data = Font(name="Calibri", size=12, bold=False)
        font_board = Font(name="Calibri", size=13, bold=True)  # Жирный бортовой

        center_alignment = Alignment(horizontal="center", vertical="center")
        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Тонкая серая сетка границ ячеек
        thin_side = Side(border_style="thin", color="CCCCCC")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # Устанавливаем высоту для ШАПКИ таблицы (первая строка)
        worksheet.row_dimensions[1].height = 26
        
        # Применяем стиль к шапке таблицы (первая строка)
        for col_idx in range(1, len(COLUMNS_EQUIPMENT) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = center_alignment

        # Применяем стиль и увеличиваем высоту строк для данных из базы
        for row_idx, row in enumerate(
            worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column,
            ),
            start=2,
        ):
            # Устанавливаем увеличенную высоту строки с данными
            worksheet.row_dimensions[row_idx].height = 22
            
            for col_idx, cell in enumerate(row, start=1):
                cell.alignment = center_alignment
                cell.border = thin_border

                # Окрашиваем каждую вторую строку данных в серый
                if row_idx % 2 == 0:
                    cell.fill = gray_fill

                # Первая колонка (Бортовой номер) — ЖИРНЫЙ шрифт
                if col_idx == 1:
                    cell.font = font_board
                else:
                    cell.font = font_data

        # 2. Добавление 10 пустых строк для заполнения от руки
        start_empty_row = worksheet.max_row + 1
        for row_idx in range(start_empty_row, start_empty_row + 60):
            # Большая высота строки, чтобы было удобно писать ручкой
            worksheet.row_dimensions[row_idx].height = 24

            for col_idx in range(1, len(COLUMNS_EQUIPMENT) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_alignment

                # Продолжаем красить полосы даже на пустых строках
                if row_idx % 2 == 0:
                    cell.fill = gray_fill

        # 3. Подготовка к печати: альбомный лист и подгонка по ширине страницы
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True

        # Применяем ширину столбцов
        column_widths = {
            "A": 22, "B": 28, "C": 22, "D": 25, 
            "E": 25, "F": 28, "G": 25, "H": 18
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

    output.seek(0)
    return output
